#!/usr/bin/env python3
"""
自動化腳本：處理 DAQ970A 資料檔案
功能：
1. 掃描所有資料夾中的 Excel 檔案
2. 提取欄位 101 (°C) 到 322 (无) 的數值
3. 忽略時間欄位
4. 計算平均值並輸出結果

使用方法：
    python3 process_data.py -p /path/to/data
"""

import openpyxl
import os
import re
import argparse
from pathlib import Path
from collections import defaultdict

def extract_channel_number(header):
    """從標題中提取通道號碼，例如 '101 (°C)' -> 101"""
    if isinstance(header, str):
        match = re.match(r'(\d+)\s*\(', header)
        if match:
            return int(match.group(1))
    return None

def is_valid_value(value):
    """檢查是否是有效的數值（排除異常值和時間戳）"""
    if value is None:
        return False
    if not isinstance(value, (int, float)):
        return False
    # 排除異常值（-9.9e+37 或其他極端值）
    if abs(value) > 1e10:
        return False
    return True

def find_header_row(ws):
    """自動找到標題行（包含通道號碼的行）"""
    for row_idx in range(1, min(ws.max_row + 1, 100)):  # 檢查前100行
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        # 檢查是否包含通道號碼（101-115）
        for cell in row:
            if isinstance(cell, str) and extract_channel_number(cell):
                return row_idx
    return None

def process_excel_file(file_path):
    """
    處理單個 Excel 檔案
    返回：{通道號: [數值列表]}
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 自動找到標題行
        header_row = find_header_row(ws)
        if header_row is None:
            return {}

        headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]

        # 找出需要的欄位（101-115，排除時間欄位）
        channel_columns = {}  # {通道號: 列索引}
        for col_idx, header in enumerate(headers, 1):
            if header and isinstance(header, str) and '时间' not in header:  # 排除時間欄位
                channel_num = extract_channel_number(header)
                if channel_num and 101 <= channel_num <= 322:
                    channel_columns[channel_num] = col_idx

        # 提取資料（從標題行之後開始）
        channel_data = defaultdict(list)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            for channel_num, col_idx in channel_columns.items():
                value = row[col_idx - 1]  # 列索引是1-based
                if is_valid_value(value):
                    channel_data[channel_num].append(value)

        return dict(channel_data)

    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return {}

def calculate_averages(channel_data):
    """計算每個通道的平均值"""
    averages = {}
    for channel_num in sorted(channel_data.keys()):
        values = channel_data[channel_num]
        if values:
            averages[channel_num] = sum(values) / len(values)
    return averages

def extract_power_value(folder_name):
    """從資料夾名稱中提取功率值用於排序，例如 '18W' -> 18"""
    import re
    match = re.match(r'(\d+)', folder_name)
    if match:
        return int(match.group(1))
    return float('inf')  # 無法解析的放在最後

def main():
    """主函數"""
    # 設定命令行參數
    parser = argparse.ArgumentParser(
        description='DAQ970A 資料處理工具 - 提取和計算通道平均值',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
範例：
  python3 process_data.py -p /Users/harry.chen/code/sstr/zhong/20230618
  python3 process_data.py -p ./data
        '''
    )

    parser.add_argument(
        '-p', '--path',
        type=str,
        required=True,
        help='資料夾路徑（必需）。應包含多個子資料夾，每個子資料夾內有 Excel 檔案'
    )

    args = parser.parse_args()

    # 驗證路徑
    base_path = Path(args.path)
    if not base_path.exists():
        print(f"❌ 錯誤：路徑不存在: {base_path}")
        return

    if not base_path.is_dir():
        print(f"❌ 錯誤：路徑不是資料夾: {base_path}")
        return

    # 收集所有資料夾
    folders = sorted([d for d in base_path.iterdir() if d.is_dir()])

    print("=" * 80)
    print("DAQ970A 資料處理結果")
    print("=" * 80)

    all_results = {}

    for folder in folders:
        file_path = folder / "1 DAQ970A USB0-0x2A8D-0x5101-MY58005988-0-INSTR.xlsx"

        if file_path.exists():
            print(f"\n處理資料夾: {folder.name}")
            print("-" * 80)

            # 處理檔案
            channel_data = process_excel_file(file_path)

            if channel_data:
                # 計算平均值
                averages = calculate_averages(channel_data)
                all_results[folder.name] = averages

                # 輸出結果
                print(f"{'通道':<10} {'平均值':<20} {'資料點數':<15}")
                print("-" * 45)
                for channel_num in sorted(averages.keys()):
                    avg = averages[channel_num]
                    count = len(channel_data[channel_num])
                    print(f"{channel_num:<10} {avg:<20.6f} {count:<15}")
            else:
                print("未找到相關資料")

    # 輸出總結
    print("\n" + "=" * 80)
    print("總結 - 所有資料夾的平均值")
    print("=" * 80)

    # 建立一個表格，每行是一個資料夾，每列是一個通道
    all_channels = set()
    for results in all_results.values():
        all_channels.update(results.keys())

    all_channels = sorted(all_channels)

    # 按功率值排序資料夾
    sorted_folders = sorted(all_results.keys(), key=extract_power_value)

    # 列印表頭
    print(f"{'資料夾':<20}", end="")
    for channel in all_channels:
        print(f"{channel:<15}", end="")
    print()
    print("-" * (20 + len(all_channels) * 15))

    # 列印資料
    for folder_name in sorted_folders:
        print(f"{folder_name:<20}", end="")
        for channel in all_channels:
            value = all_results[folder_name].get(channel, 0)
            print(f"{value:<15.6f}", end="")
        print()

    # 保存到 CSV 檔案
    import csv
    csv_path = base_path.parent / "results.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        # 寫入表頭
        writer.writerow(['資料夾'] + all_channels)
        # 寫入資料（按功率排序）
        for folder_name in sorted_folders:
            row = [folder_name]
            for channel in all_channels:
                value = all_results[folder_name].get(channel, 0)
                row.append(f"{value:.6f}")
            writer.writerow(row)

    print(f"\n結果已保存到: {csv_path}")

if __name__ == "__main__":
    main()
